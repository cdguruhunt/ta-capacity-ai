import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel_report(inputs, result) -> bytes:
    """Generate a formatted Excel workforce planning report."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TA Structure Report"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    subheader_fill = PatternFill("solid", fgColor="533483")
    subheader_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    label_font = Font(name="Calibri", bold=True, size=10)
    value_font = Font(name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "TalentScale AI — Workforce Planning Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1A1A2E")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="888888")
    ws["A2"].alignment = center

    # Inputs section
    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "INPUT PARAMETERS"
    ws[f"A{row}"].font = subheader_font
    ws[f"A{row}"].fill = subheader_fill
    ws[f"A{row}"].alignment = center
    row += 1

    input_data = [
        ("Hiring Target", inputs.hiring_target),
        ("Company Type", inputs.company_type.title()),
        ("Dropout Ratio", f"{inputs.dropout_ratio}%"),
        ("Complexity Factor", inputs.complexity_factor.title()),
        ("Recruiter Productivity", f"{inputs.recruiter_productivity} hires/yr"),
        ("Geography", inputs.geography.title()),
        ("AI / Niche Hiring", f"{inputs.ai_niche_percent}%"),
    ]

    for label, val in input_data:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = str(val)
        ws[f"B{row}"].font = value_font
        for col in ["A", "B"]:
            ws[f"{col}{row}"].border = border
        row += 1

    # Results section
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "RECOMMENDED TA STRUCTURE"
    ws[f"A{row}"].font = subheader_font
    ws[f"A{row}"].fill = subheader_fill
    ws[f"A{row}"].alignment = center
    row += 1

    result_data = [
        ("Adjusted Hiring Target", result.adjusted_hiring),
        ("Total TA Team", result.total_ta),
        ("", ""),
        ("Junior Recruiters (40%)", result.junior_recruiters),
        ("Recruiters (35%)", result.recruiters),
        ("Senior Recruiters (25%)", result.senior_recruiters),
        ("Sourcers", result.sourcers),
        ("Coordinators", result.coordinators),
        ("Team Leads", result.leads),
        ("Managers", result.managers),
        ("TA Head", result.ta_head),
        ("", ""),
        ("Hiring Capacity", result.hiring_capacity),
        ("Estimated Timeline", result.timeline),
        ("Recruiter Utilization", f"{result.utilization}%"),
        ("Estimated TA Cost (Annual)", result.estimated_cost),
        ("Cost Per Hire", result.cost_per_hire),
        ("Model Used", result.model_used),
    ]

    for label, val in result_data:
        if label == "":
            row += 1
            continue
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = label_font
        ws[f"B{row}"] = str(val)
        ws[f"B{row}"].font = value_font
        for col in ["A", "B"]:
            ws[f"{col}{row}"].border = border
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
